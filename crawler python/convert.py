import openpyxl
import json

excelPath = 'C:/Users/56thf/Desktop/project/crawler python/PopUpStore.xlsx'
jsonPath = 'C:/Users/56thf/Desktop/project/crawler python/PopUpStore.json'

# 기존 엑셀 파일 불러오기
wb = openpyxl.load_workbook(excelPath, read_only = True)

sheet = wb.worksheets[0]

# 상단을 키 값
keyList = []
keyIndex = 0
for col_num in range(1, sheet.max_column + 1): # A ~
    keyList.append(sheet.cell(row=1, column=col_num).value)

print(keyList)

data_dict = {}
for row_num in range(2, sheet.max_row + 1): # 두 번째 행부터
    tmp_dict = {}
    for col_num in range(1, sheet.max_column + 1):
        val = sheet.cell(row = row_num, column = col_num).value
        tmp_dict[keyList[col_num - 1]] = val

    data_dict[tmp_dict[keyList[keyIndex]]] = tmp_dict

data_dict

print(data_dict)

# 엑셀 닫기
wb.close()

with open(jsonPath, 'w', encoding='utf-8') as fp:
    json.dump(data_dict, fp, indent=4, ensure_ascii=False)